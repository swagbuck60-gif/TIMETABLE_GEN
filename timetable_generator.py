import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
from datetime import datetime

# Page config for GitHub/Streamlit Cloud
st.set_page_config(
    page_title="🎓 School Timetable Generator",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="expanded"
)

class StreamlitTimetableGenerator:
    COLORS = {
        'school': '2E8B57', 'class': '32CD32', 'teacher': 'FFD700',
        'day': '1E90FF', 'period': '4169E1', 'data': 'E0F2F1'
    }
    DAYS = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT']
    PERIODS = ['1', '2', '3', '4', '5', '6', '7', '8']

    def analyze_file(self, df):
        """Auto-detect file format"""
        data_start = 0
        for i, row in df.iterrows():
            if pd.notna(row.iloc[1]) and str(row.iloc[1]).strip() and 'NAME' not in str(row.iloc[1]).upper():
                data_start = i
                break
        df = df.iloc[data_start:].reset_index(drop=True)
        
        has_subject = False
        if len(df.columns) > 3 and pd.notna(df.iloc[0, 3]):
            sample_subject = str(df.iloc[0, 3]).strip().upper()
            if sample_subject in ['MATHS', 'CHEM', 'BIO', 'ENG', 'PHY', 'HINDI', 'CS', 'HIST']:
                has_subject = True
        
        return df, has_subject

    def parse_data(self, df, has_subject):
        """Extract teachers and classes"""
        teachers = []
        classes = set()
        
        for idx, row in df.iterrows():
            name = str(row.iloc[1]).strip()
            if not name or pd.isna(row.iloc[1]):
                continue
                
            teacher_data = {
                'name': name,
                'designation': str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else '',
                'subject': str(row.iloc[3]).strip() if has_subject and pd.notna(row.iloc[3]) else name,
                'periods': row.iloc[4:-1].dropna().tolist() if len(row) > 4 else []
            }
            
            for cell in teacher_data['periods']:
                if pd.notna(cell) and isinstance(cell, str):
                    class_name = str(cell).strip()
                    if class_name and len(class_name) > 2:
                        classes.add(class_name)
            
            teachers.append(teacher_data)
        
        return teachers, sorted(list(classes))

    def style_cell(self, cell, fill_color, font_size=11, bold=False, text_color='000000'):
        cell.font = Font(bold=bold, size=font_size, color=text_color)
        cell.fill = PatternFill(start_color=self.COLORS[fill_color], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = border

    def create_workbook(self, teachers, classes, school_name):
        """Create Excel workbook"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Teacher Sheet
        ws = wb.create_sheet('👨‍🏫 Teacher Timetable', 0)
        ws.column_dimensions['A'].width = 20
        for col in range(2, 10): ws.column_dimensions[get_column_letter(col)].width = 13
        
        row = 1
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = f"🏫 {school_name}"
        self.style_cell(cell, 'school', 16, True, 'FFFFFF')
        ws.row_dimensions[row].height = 35
        row += 2
        
        for teacher in teachers:
            ws.merge_cells(f'A{row}:I{row}')
            cell = ws[f'A{row}']
            cell.value = f"{teacher['name']}\n({teacher['designation']})\n📖 {teacher['subject']}"
            self.style_cell(cell, 'teacher', 12, True)
            ws.row_dimensions[row].height = 45
            row += 2
            
            # Headers
            ws[f'A{row}'] = 'Day/Period'
            self.style_cell(ws[f'A{row}'], 'period', 11, True, 'FFFFFF')
            for p_idx in range(8):
                cell = ws.cell(row, p_idx + 2, f'P{self.PERIODS[p_idx]}')
                self.style_cell(cell, 'period', 10, True, 'FFFFFF')
            row += 1
            
            # Data
            periods = teacher['periods']
            for d_idx, day in enumerate(self.DAYS):
                cell = ws.cell(row, 1, day)
                self.style_cell(cell, 'day', 11, True, 'FFFFFF')
                for p_idx in range(8):
                    col_idx = d_idx * 8 + p_idx + 4
                    cell = ws.cell(row, p_idx + 2)
                    if col_idx < len(periods) and pd.notna(periods[col_idx]):
                        cell.value = str(periods[col_idx]).strip()
                    self.style_cell(cell, 'data')
                ws.row_dimensions[row].height = 22
                row += 1
            row += 2
        
        # Class Sheet
        ws = wb.create_sheet('📚 Class Timetable', 1)
        ws.column_dimensions['A'].width = 18
        for col in range(2, 10): ws.column_dimensions[get_column_letter(col)].width = 14
        
        row = 1
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws[f'A{row}']
        cell.value = f"🏫 {school_name}"
        self.style_cell(cell, 'school', 16, True, 'FFFFFF')
        ws.row_dimensions[row].height = 35
        row += 2
        
        for cls in classes:
            ws.merge_cells(f'A{row}:I{row}')
            cell = ws[f'A{row}']
            cell.value = f"📖 Class {cls}"
            self.style_cell(cell, 'class', 13, True, 'FFFFFF')
            ws.row_dimensions[row].height = 30
            row += 2
            
            ws[f'A{row}'] = 'Day/Period'
            self.style_cell(ws[f'A{row}'], 'period', 11, True, 'FFFFFF')
            for p_idx in range(8):
                cell = ws.cell(row, p_idx + 2, f'P{self.PERIODS[p_idx]}')
                self.style_cell(cell, 'period', 10, True, 'FFFFFF')
            row += 1
            
            for d_idx, day in enumerate(self.DAYS):
                cell = ws.cell(row, 1, day)
                self.style_cell(cell, 'day', 11, True, 'FFFFFF')
                for p_idx in range(8):
                    cell = ws.cell(row, p_idx + 2)
                    subjects = []
                    for teacher in teachers:
                        periods = teacher['periods']
                        col_idx = d_idx * 8 + p_idx + 4
                        if col_idx < len(periods) and pd.notna(periods[col_idx]):
                            if str(periods[col_idx]).strip() == cls:
                                subjects.append(teacher['subject'])
                    if subjects:
                        cell.value = '/'.join(subjects)
                    self.style_cell(cell, 'data')
                ws.row_dimensions[row].height = 25
                row += 1
            row += 2
        
        return wb

def main():
    st.title("🎓 School Timetable Generator")
    st.markdown("**Upload your Excel file → Get colorful timetables instantly!** ✨")
    
    # Sidebar
    st.sidebar.header("⚙️ Settings")
    school_name = st.sidebar.text_input("School Name", "Jawahar Navodaya Vidyalaya")
    
    # File uploader
    uploaded_file = st.file_uploader(
        "📁 Upload your timetable Excel file",
        type=['xlsx', 'xls'],
        help="Upload file with 'SCHOOL TIMETABLE' sheet (like your sample)"
    )
    
    if uploaded_file is not None:
        # Read file
        df = pd.read_excel(uploaded_file, sheet_name='SCHOOL TIMETABLE')
        st.success(f"✅ Loaded {len(df)} rows!")
        
        with st.spinner("🔍 Analyzing file structure..."):
            gen = StreamlitTimetableGenerator()
            df_clean, has_subject = gen.analyze_file(df)
            teachers, classes = gen.parse_data(df_clean, has_subject)
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("👨‍🏫 Teachers", len(teachers))
        with col2:
            st.metric("📚 Classes", len(classes))
        with col3:
            st.metric("🎨 Colors", "7 Distinct")
        
        st.success(f"""
        **File Analysis Complete!**
        - Teachers found: {len(teachers)}
        - Classes found: {len(classes)} ({', '.join(classes[:5])}{'...' if len(classes)>5 else ''})
        - Subject column: {'✅ Yes' if has_subject else '❌ No (auto-detected)'}
        """)
        
        if st.button("🚀 GENERATE TIMETABLE", type="primary"):
            with st.spinner("🎨 Creating beautiful Excel file..."):
                wb = gen.create_workbook(teachers, classes, school_name)
                
                # Save to bytes
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                # Download button
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Timetable_{school_name.replace(' ', '_')}_{timestamp}.xlsx"
                
                st.balloons()
                st.success("✨ Timetable generated successfully!")
                
                st.download_button(
                    label="📥 Download Timetable.xlsx",
                    data=output.getvalue(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.info("👆 Upload your Excel file to get started!")
        st.markdown("---")
        st.markdown("""
        ## ✨ **Features**
        - 📱 **Drag & drop** file upload
        - 🔍 **Auto-detects** file format
        - 🌈 **7 beautiful colors** (School/Class/Teacher/Day/Period/Data)
        - 📊 **Periods in COLUMNS**, Days in **ROWS**
        - 👨‍🏫 **Teacher sheet** with subjects
        - 📚 **Class sheet** shows SUBJECTS (MATHS/ENG not teachers)
        - 🎨 **Professional borders** & formatting
        - ⚡ **Instant download**
        """)

if __name__ == "__main__":
    main()
